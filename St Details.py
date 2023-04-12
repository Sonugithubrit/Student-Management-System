from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
import mysql.connector
from tkinter import messagebox
import cv2
import os
import csv
from tkinter import filedialog
from datetime import date
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib

background="#06283D"
framebg="#EDEDED"
framefg="#06283D"


root=Tk()
root.title("Star x.prt computer")
root.geometry("1367x700+0+0")
root.config(bg=background)
root.wm_iconbitmap("icon.ico")

file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Student Name"
    sheet['C1']="Father Name"
    sheet['D1']="Mother Name"
    sheet['E1']="Date of Birth"
    sheet['F1']="Gender"
    sheet['G1']="Religion"
    sheet['H1']="Categary"
    sheet['I1']="Adhaar N0."
    sheet['J1']="Phone No."
    sheet['K1']="Village"
    sheet['L1']="Post"
    sheet['M1']="Dist"
    sheet['N1']="State"
    sheet['O1']="Pin No."
    sheet['P1']="Addmission Data"
    sheet['Q1']="Addmission Fee"
    sheet['R1']="Course"
    sheet['S1']="Course Fee"
    sheet['T1']="Batch Time"
    sheet['U1']="Qualification"
    sheet['V1']="Total fee submit"

    
    file.save('Student_data.xlsx')

# Exit
def Exit():
    root.destroy()


# ======================ShowImage======================
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",filetype=(("JPG File",".jpg"),
                                                                            ("PNG File",".png"),
                                                                            ("JPEG File",".jpeg"),
                                                                            ("All files",".txt")))
    img=(Image.open(filename))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2


# ===================================Registration No.===============================
def registration_no():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value
    # print(max_row_value+1)

    try:
        Registration.set(max_row_value+1)

    except:
        Registration.set("1")

# ===========================Clear===========================
def Clear():
    global img
    Name.set('')
    Fname.set('')
    Mname.set('')
    radio.set('')
    Religion.set('')
    Categary.set('')
    DOB.set('')
    Adhar.set('')
    Phone.set('')
    Village.set('')
    post.set('')
    Dist.set('')
    State.set('')
    Pin.set('')
    Addmid.set('')
    Addmif.set('')
    Course.set('')
    Coursef.set('')
    BatchT.set('')
    Quali.set('')
    Total.set('')

    registration_no()

    savebutton.config(state='normal')

    img1=PhotoImage(file='imges/logo1.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=""

# ==========================Sab=ve===========================
def Save():
    R1=Registration.get()
    N1=Name.get()
    C1=Course.get()
    try:
        G1=gender
    except:
        messagebox.showerror("Error","Select Gender!")

    F1=Fname.get()
    M1=Mname.get()
    R3=Religion.get()
    C2=Categary.get()
    D1=DOB.get()
    A1=Adhar.get()
    P1=Phone.get()
    V1=Village.get()
    P2=post.get()
    D2=Dist.get()
    S1=State.get()
    P3=Pin.get()
    A2=Addmid.get()
    A3=Addmif.get()
    C3=Coursef.get()
    B1=BatchT.get()
    Q1=Quali.get()
    T1=Total.get()
    
    if N1=="" or C1=="" or F1=="" or M1=="" or R3=="" or C2=="" or D1=="" or A1=="" or P1=="" or V1=="" or P2=="" or D2=="" or S1=="" or P3=="" or A2=="" or A3=="" or C3=="" or B1=="" or Q1=="" or T1=="":
        messagebox.showerror("error","Few Data is missing!")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=F1)
        sheet.cell(column=4,row=sheet.max_row,value=M1)
        sheet.cell(column=5,row=sheet.max_row,value=D1)
        sheet.cell(column=6,row=sheet.max_row,value=G1)
        sheet.cell(column=7,row=sheet.max_row,value=R3)
        sheet.cell(column=8,row=sheet.max_row,value=C2)
        sheet.cell(column=9,row=sheet.max_row,value=A1)
        sheet.cell(column=10,row=sheet.max_row,value=P1)
        sheet.cell(column=11,row=sheet.max_row,value=V1)
        sheet.cell(column=12,row=sheet.max_row,value=P2)
        sheet.cell(column=13,row=sheet.max_row,value=D2)
        sheet.cell(column=14,row=sheet.max_row,value=S1)
        sheet.cell(column=15,row=sheet.max_row,value=P3)
        sheet.cell(column=16,row=sheet.max_row,value=A2)
        sheet.cell(column=17,row=sheet.max_row,value=A3)
        sheet.cell(column=18,row=sheet.max_row,value=C1)
        sheet.cell(column=19,row=sheet.max_row,value=C3)
        sheet.cell(column=20,row=sheet.max_row,value=B1)
        sheet.cell(column=21,row=sheet.max_row,value=Q1)
        sheet.cell(column=22,row=sheet.max_row,value=T1)

        file.save(r'Student_data.xlsx')

        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile Picture is not available!!!!")
        messagebox.showinfo("info","Sucessfully data Entered!!!")

        Clear()
        registration_no()

# ==================Search==============================
def search():
    text=Search.get()
    Clear()
    savebutton.config(state='disable')

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active
    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
            # print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

            # print(reg_no_position)
            # print(reg_number)

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalied","Invalied registration number!!!!")
    
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value
    x13=sheet.cell(row=int(reg_number),column=13).value
    x14=sheet.cell(row=int(reg_number),column=14).value
    x15=sheet.cell(row=int(reg_number),column=15).value
    x16=sheet.cell(row=int(reg_number),column=16).value
    x17=sheet.cell(row=int(reg_number),column=17).value
    x18=sheet.cell(row=int(reg_number),column=18).value
    x19=sheet.cell(row=int(reg_number),column=19).value
    x20=sheet.cell(row=int(reg_number),column=20).value
    x21=sheet.cell(row=int(reg_number),column=21).value
    x22=sheet.cell(row=int(reg_number),column=22).value
    # print(x1)
    # print(x2)
    # print(x3)
    # print(x4)
    # print(x5)
    # print(x6)
    # print(x7)
    # print(x8)
    # print(x9)
    # print(x10)
    # print(x11)
    # print(x12)
    # print(x13)
    # print(x14)
    # print(x15)
    # print(x16)
    # print(x17)
    # print(x18)
    # print(x19)
    # print(x20)
    # print(x21)

    Registration.set(x1)
    Name.set(x2)
    Course.set(x18)
    if x6=='Female':
        R2.select()
    else:
        R1.select()
    Fname.set(x3)
    Mname.set(x4)
    # radio.set('x5')
    Religion.set(x7)
    Categary.set(x8)
    DOB.set(x5)
    Adhar.set(x9)
    Phone.set(x10)
    Village.set(x11)
    post.set(x12)
    Dist.set(x13)
    State.set(x14)
    Pin.set(x15)
    Addmid.set(x16)
    Addmif.set(x17)
    Coursef.set(x19)
    BatchT.set(x20)
    Quali.set(x21)
    Total.set(x22)
    img = (Image.open("Student Images/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

# ====================Update============================
def Update():
    R1=Registration.get()
    N1=Name.get()
    C1=Course.get()
    selection()
    G1=gender
    F1=Fname.get()
    M1=Mname.get()
    R3=Religion.get()
    C2=Categary.get()
    D1=DOB.get()
    A1=Adhar.get()
    P1=Phone.get()
    V1=Village.get()
    P2=post.get()
    D2=Dist.get()
    S1=State.get()
    P3=Pin.get()
    A2=Addmid.get()
    A3=Addmif.get()
    C3=Coursef.get()
    B1=BatchT.get()
    Q1=Quali.get()
    T1=Total.get()

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

            print(reg_number)

    sheet.cell(column=1,row=int(reg_number),value=R1)
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=F1)
    sheet.cell(column=4,row=int(reg_number),value=M1)
    sheet.cell(column=5,row=int(reg_number),value=D1)
    sheet.cell(column=6,row=int(reg_number),value=G1)
    sheet.cell(column=7,row=int(reg_number),value=R3)
    sheet.cell(column=8,row=int(reg_number),value=C2)
    sheet.cell(column=9,row=int(reg_number),value=A1)
    sheet.cell(column=10,row=int(reg_number),value=P1)
    sheet.cell(column=11,row=int(reg_number),value=V1)
    sheet.cell(column=12,row=int(reg_number),value=P1)
    sheet.cell(column=13,row=int(reg_number),value=D2)
    sheet.cell(column=14,row=int(reg_number),value=S1)
    sheet.cell(column=15,row=int(reg_number),value=P2)
    sheet.cell(column=16,row=int(reg_number),value=A2)
    sheet.cell(column=17,row=int(reg_number),value=A3)
    sheet.cell(column=18,row=int(reg_number),value=C1)
    sheet.cell(column=19,row=int(reg_number),value=C2)
    sheet.cell(column=20,row=int(reg_number),value=B1)
    sheet.cell(column=21,row=int(reg_number),value=Q1)
    sheet.cell(column=22,row=int(reg_number),value=T1)

    file.save(r"Student_data.xlsx")


    

#gender

def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
        print(gender)
    else:
        gender="Female"
        print(gender)

# Top Frame
Label(root,text="Email: starxprtcomputer2023@gmail.com",width=10,height=3,bg="#f0687c",fg="blue",anchor='e').pack(side=TOP,fill=X)
Label(root,text="Star X-Prt Computer",width=10,bg="#c36464",fg="#fff",font='arial 40 bold').pack(side=TOP,fill=X)
Label(root,text="Technical Institute Dhanghata S.K.N",width=115,bg="#c36464",fg="#fff",font='arial 15 bold').place(x=0,y=110)
# Search Box to Update
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font='arial 20').place(x=980,y=150)
search=Button(root,text="Search",width=10,cursor="hand2",bg="#68ddfa",font="arial 13 bold",command=search)
search.place(x=1230,y=150)

# img_logo=Image.open(r'imges/logo1.png')
# img_logo=img_logo.resize((75,75),Image.ANTIALIAS)
# root.photo_logo=ImageTk.PhotoImage(img_logo)

# root.logo=Label(root,image=root.photo_logo,bg='#c36464')
# root.logo.place(x=280,y=58,width=75,height=75)

# img_logo1=Image.open(r'imges/payment.jfif')
# img_logo1=img_logo1.resize((88,88),Image.ANTIALIAS)
# root.photo_logo1=ImageTk.PhotoImage(img_logo1)

# root.logo1=Label(root,image=root.photo_logo1,bg='#c36464')
# root.logo1.place(x=1230,y=52,width=88,height=88)
# Registration Date
Label(root,text="Registration No.",font="arial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date:",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)

Registration=IntVar()
Date=StringVar()

reg_entry=Entry(root,textvariable=Registration,width=15,font='arial 10')
reg_entry.place(x=160,y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%y")
date_entry =Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)

Date.set(d1)

# Student details

obg=LabelFrame(root,text="Student's Details",font=20,bd=2,width=1000,height=250,bg=framebg,fg=framefg,relief=GROOVE)
obg.place(x=30,y=200)

Name=StringVar()
Label(obg,text="Student Name",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=10)
name_entry=Entry(obg,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=200,y=14)

Fname=StringVar()
Label(obg,text="Father's Name",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=40)
fname_entry=Entry(obg,textvariable=Fname,width=20,font="arial 10")
fname_entry.place(x=200,y=44)


Mname=StringVar()
Label(obg,text="Mother's Name",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=70)
mname_entry=Entry(obg,textvariable=Mname,width=20,font="arial 10")
mname_entry.place(x=200,y=74)

# Gender
radio=IntVar()
lbl_gen=Label(obg,text='Gender',font="arial 13",bg=framebg,fg=framefg)
lbl_gen.place(x=30,y=100)

R1 = Radiobutton(obg,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=200,y=104)

R2 = Radiobutton(obg,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=250,y=104)

# Religion
Religion=StringVar()
lbl_rel=Label(obg,text='Religion',font="arial 13",bg=framebg,fg=framefg)
lbl_rel.place(x=30,y=130)

combo_gen=ttk.Combobox(obg,textvariable=Religion,font=('arial,11'),width=13,state='readonly')
combo_gen['value']=('Select Religion','Hindu','Muslim','Sikh','Cristian','Other')
combo_gen.current(0)
combo_gen.place(x=200,y=134)


 # Categary
Categary=StringVar()
lbl_cat=Label(obg,text='Categary',font="arial 13",bg=framebg,fg=framefg)
lbl_cat.place(x=30,y=160)

combo_gen=ttk.Combobox(obg,textvariable=Categary,font=('arial,11'),width=13,state='readonly')
combo_gen['value']=('Select Categary','Gen','OBC','Sc',"St")
combo_gen.current(0)
combo_gen.place(x=200,y=164)

# DOB
DOB=StringVar()
lbl_dob=Label(obg,text='Date Of Birth',font="arial 13",bg=framebg,fg=framefg)
lbl_dob.place(x=30,y=190)
name_entry=Entry(obg,textvariable=DOB,width=20,font="arial 10")
name_entry.place(x=200,y=194)

# txt_dob=ttk.Entry(upper_frame,textvariable=self.var_dob,width=22,font=('arial',11))
# txt_dob.grid(row=6,column=1,padx=2,pady=4)

Adhar=StringVar()
Label(obg,text="Adhaar No. ",font="arial 13",bg=framebg,fg=framefg).place(x=600,y=10)
name_entry=Entry(obg,textvariable=Adhar,width=20,font="arial 10")
name_entry.place(x=700,y=14)

Phone=StringVar()
Label(obg,text="Phone No.",font="arial 13",bg=framebg,fg=framefg).place(x=600,y=40)
name_entry=Entry(obg,textvariable=Phone,width=20,font="arial 10")
name_entry.place(x=700,y=44)

Village=StringVar()
Label(obg,text="Village",font="arial 13",bg=framebg,fg=framefg).place(x=600,y=70)
name_entry=Entry(obg,textvariable=Village,width=20,font="arial 10")
name_entry.place(x=700,y=74)

post=StringVar()
Label(obg,text="Post",font="arial 13",bg=framebg,fg=framefg).place(x=600,y=100)
name_entry=Entry(obg,textvariable=post,width=20,font="arial 10")
name_entry.place(x=700,y=104)

Dist=StringVar()
Label(obg,text="District",font="arial 13",bg=framebg,fg=framefg).place(x=600,y=130)
name_entry=Entry(obg,textvariable=Dist,width=20,font="arial 10")
name_entry.place(x=700,y=134)

State=StringVar()
Label(obg,text="State",font="arial 13",bg=framebg,fg=framefg).place(x=600,y=160)
name_entry=Entry(obg,textvariable=State,width=20,font="arial 10")
name_entry.place(x=700,y=164)

Pin=StringVar()
Label(obg,text="Pin No.",font="arial 13",bg=framebg,fg=framefg).place(x=600,y=190)
name_entry=Entry(obg,textvariable=Pin,width=20,font="arial 10")
name_entry.place(x=700,y=194)

# Course details

obj=LabelFrame(root,text="Course Details",font=20,bd=2,width=1000,height=230,bg=framebg,fg=framefg,relief=GROOVE)
obj.place(x=30,y=460)

Addmid=StringVar()
Label(obj,text="Addmission Date",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=30)
name_entry=Entry(obj,textvariable=Addmid,width=20,font="arial 10")
name_entry.place(x=200,y=34)

Addmif=StringVar()
Label(obj,text="Addmission Fee",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=70)
name_entry=Entry(obj,textvariable=Addmif,width=20,font="arial 10")
name_entry.place(x=200,y=74)

# Course
Course=StringVar()
lbl_cou=Label(obj,text='Course',font="arial 13",bg=framebg,fg=framefg)
lbl_cou.place(x=30,y=110)

combo_cou=ttk.Combobox(obj,textvariable=Course,font=('arial,11'),width=13,state='readonly')
combo_cou['value']=('Select Course','CCC','DCA','DCH',"ADCA",'PGDCA','BCC','DTP','CTT','DFA','DAHN','MSIT','O Level','A Level','Internet','Networking','Typing','Tally prime','Tally Marge','Graphic Designing','2D Animation','Web Designing','C Programming','C++','HTML','CSS','JavaScript','Python')
combo_cou.current(0)
combo_cou.place(x=200,y=114)

# Course Fee
Coursef=StringVar()
lbl_couf=Label(obj,text='Course Fee p/m',font="arial 13",bg=framebg,fg=framefg)
lbl_couf.place(x=30,y=150)

combo_couf=ttk.Combobox(obj,textvariable=Coursef,font=('arial,11'),width=13,state='readonly')
combo_couf['value']=('Select Course Fee','CCC- 600','DCA- 600','DCH- 900',"ADCA- 600",'PGDCA- 1000','BCC- 700','DTP- 600','CTT- 1200','DFA- 600','DAHN- 1200','MSIT- 700','O Level- 1000','A Level- 1000','Internet- 700','Networking- 1000','Typing- 400','Tally prime- 700','Tally Marge- 700','Graphic Designing- 700','2D Animation- 1000','Web Designing-1000','C Programming- 1000','C++ -1000','HTML- 1000','CSS- 1000','JavaScript- 1000','Python- 1000')
combo_couf.current(0)
combo_couf.place(x=200,y=154)

# Batch Time
BatchT=StringVar()
lbl_batch=Label(obj,text='Batch Time',font="arial 13",bg=framebg,fg=framefg)
lbl_batch.place(x=600,y=45)

combo_batch=ttk.Combobox(obj,textvariable=BatchT,font=('arial,11'),width=13,state='readonly')
combo_batch['value']=('Select Batch Time','7am To 9am','9am To 11','11am To 1pm',"1pm To 3pm",'3pm To 5pm')
combo_batch.current(0)
combo_batch.place(x=800,y=49)

# Qualification
Quali=StringVar()
lbl_qualification=Label(obj,text='Qualifcation',font="arial 13",bg=framebg,fg=framefg)
lbl_qualification.place(x=600,y=90)


combo_qualification=ttk.Combobox(obj,textvariable=Quali,font=('arial,11'),width=13,state='readonly')
combo_qualification['value']=('Select Qualification','High School','Intermediate','Graduation',"Other")
combo_qualification.current(0)

combo_qualification.place(x=800,y=94)
# Total Fee submit
Total=StringVar()
lbl_dob=Label(obj,text='Total Fee Submit',font="arial 13",bg=framebg,fg=framefg)
lbl_dob.place(x=600,y=135)
name_entry=Entry(obj,textvariable=Total,width=20,font="arial 10")
name_entry.place(x=800,y=137)

#Image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1100,y=200)

# img_log=Image.open('img/payment.jfif')
# img_log=img_log.resize((200,200),Image.ANTIALIAS)
# root.photo_log=ImageTk.PhotoImage(img_log)

# lbl=Button(root,image=root.photo_log,bg='#c36464')
# lbl.place(x=1100,y=200,width=200,height=200)

img=PhotoImage("imges/logo1.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#Button
uploadbutton=Button(root,text="Upload",width=19,cursor="hand2",height=1,font="arial 12 bold",bg="lightblue",command=showimage)
uploadbutton.place(x=1100,y=390)

savebutton=Button(root,text="Submit",width=19,height=1,cursor="hand2",font="arial 12 bold",bg="lightgreen",command=Save)
savebutton.place(x=1100,y=450)

updatebutton=Button(root,text="Upadte",width=19,height=1,cursor="hand2",font="arial 12 bold",bg="red",command=Update)
updatebutton.place(x=1100,y=510)

resetbutton=Button(root,text="Reset",width=19,height=1,cursor="hand2",font="arial 12 bold",bg="lightpink",command=Clear)
resetbutton.place(x=1100,y=570)

exitbutton=Button(root,text="Exit",width=19,height=1,cursor="hand2",font="arial 12 bold",bg="gray",command=Exit)
exitbutton.place(x=1100,y=620)


root.mainloop()